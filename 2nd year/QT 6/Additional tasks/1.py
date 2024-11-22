import csv
import openpyxl

data = openpyxl.load_workbook(filename='data.xlsx', data_only=True)
sh_o = data.active
titles = [sh_o.cell(1, j).value for j in range(1, sh_o.max_column + 1)]
values = [[sh_o.cell(i, j).value for j in range(1, sh_o.max_column + 1)] for i in range(2, sh_o.max_row + 1)]
with open('output.csv', 'w', encoding='utf-8') as f:
    writer = csv.writer(f, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(titles)
    writer.writerows(values)